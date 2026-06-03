# data/repositorio_pedidos.py
import openpyxl
from typing import List
from pathlib import Path
from core.modelos import Pedido
from datetime import datetime
import os
import sys

class RepositorioPedidos:
    """Gestiona la lectura y persistencia de pedidos desde Excel"""
    
    def __init__(self, ruta_archivo: str = "reporte_pedidos.xlsx"):
        self.ruta_archivo = Path(self._get_root_dir() / ruta_archivo)
        
    def _get_root_dir(self) -> Path:
        """Obtiene el directorio raíz del proyecto"""
        if getattr(sys, 'frozen', False):
            return Path(os.path.dirname(os.path.abspath(sys.executable)))
        return Path(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        
    def cargar_pedidos(self) -> List[Pedido]:
        """Carga los pedidos desde el archivo Excel en orden inverso (nuevos primero)"""
        if not self.ruta_archivo.exists():
            return []
        
        try:
            wb = openpyxl.load_workbook(self.ruta_archivo, read_only=True, data_only=True)
            sheet = wb.active
            
            headers = []
            # Obtener cabecera
            for row in sheet.iter_rows(max_row=1, values_only=True):
                headers = list(row)
                break
                
            header_map = {name: idx for idx, name in enumerate(headers) if name is not None}
            
            pedidos = []
            for row_cells in sheet.iter_rows(min_row=2, values_only=True):
                factura_idx = header_map.get('Factura')
                if factura_idx is None or factura_idx >= len(row_cells) or row_cells[factura_idx] is None:
                    continue
                    
                factura = str(row_cells[factura_idx])
                
                estado_idx = header_map.get('EstadoTransaccion')
                estado = str(row_cells[estado_idx]) if (estado_idx is not None and estado_idx < len(row_cells) and row_cells[estado_idx] is not None) else ""
                
                total_idx = header_map.get('Total')
                total = 0.0
                if total_idx is not None and total_idx < len(row_cells) and row_cells[total_idx] is not None:
                    try:
                        total = float(row_cells[total_idx])
                    except (ValueError, TypeError):
                        total = 0.0
                        
                cliente_idx = header_map.get('Cliente')
                cliente = str(row_cells[cliente_idx]) if (cliente_idx is not None and cliente_idx < len(row_cells) and row_cells[cliente_idx] is not None) else ""
                
                vendedor_idx = header_map.get('Vendedor')
                vendedor = str(row_cells[vendedor_idx]) if (vendedor_idx is not None and vendedor_idx < len(row_cells) and row_cells[vendedor_idx] is not None) else ""
                
                moneda_idx = header_map.get('Moneda')
                moneda = str(row_cells[moneda_idx]) if (moneda_idx is not None and moneda_idx < len(row_cells) and row_cells[moneda_idx] is not None) else ""
                
                pedido = Pedido(
                    numero_pedido=factura,
                    fecha=datetime.now(),
                    estado=estado,
                    total=total,
                    cliente=cliente,
                    vendedor=vendedor,
                    moneda=moneda,
                )
                pedidos.append(pedido)
                
            wb.close()
            
            # INVERTIR EL ORDEN: los pedidos más nuevos (últimos en el Excel) aparecerán primero
            pedidos.reverse()
            return pedidos
        except Exception as e:
            print(f"Error al cargar pedidos del repositorio: {e}")
            return []
    
    def guardar_pedidos(self, pedidos: List[Pedido]) -> None:
        """Guarda los pedidos en el archivo Excel"""
        try:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = "Pedidos"
            
            headers = ["Factura", "EstadoTransaccion", "Total", "Cliente", "Vendedor", "Moneda"]
            sheet.append(headers)
            
            for p in pedidos:
                sheet.append([
                    p.numero_pedido,
                    p.estado,
                    p.total,
                    p.cliente,
                    p.vendedor,
                    p.moneda
                ])
                
            wb.save(self.ruta_archivo)
            wb.close()
        except Exception as e:
            print(f"Error al guardar pedidos en el repositorio: {e}")
    
    def existe_archivo(self) -> bool:
        """Verifica si el archivo existe"""
        return self.ruta_archivo.exists()