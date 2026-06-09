# data/servicio_nodo.py
from playwright.sync_api import sync_playwright
from typing import Callable, Optional, Dict, Any
from threading import Event
from core.modelos import Pedido, Credenciales
import openpyxl
import os
import sys
import tempfile
import re
from datetime import datetime, timedelta

# Fecha base fija para ordenamiento cuando el Excel no tiene columna Fecha.
# Pedidos con índice mayor → fecha más reciente → aparecen primero en la UI.
_FECHA_BASE = datetime(2000, 1, 1)


class ServicioNodo:
    """Conexión con el sistema NODO para sincronizar pedidos"""

    def __init__(self, credenciales: Credenciales, log_fn: Optional[Callable] = None):
        self.credenciales = credenciales
        self.log = log_fn or print
        self._stop_event = Event()
        self._reset_event = Event()

    def _login_nodo(self, page) -> bool:
        """Autentica en NODO y valida si las credenciales son correctas."""
        try:
            self.log("[AUTH] Autenticando en NODO...")
            if not page.query_selector("#Correo"):
                page.goto(self.credenciales.servidor, timeout=45000)
            page.wait_for_selector("#Correo", timeout=5000)
            page.fill("#Correo", self.credenciales.usuario)
            page.fill("#txtClave", self.credenciales.password)
            page.evaluate("document.getElementById('chkMantenerSesion').click()")
            page.click("button[type='submit']")

            try:
                page.wait_for_url(re.compile(r".*(Home|LibroVenta).*"), timeout=20000, wait_until="commit")
                self.log("[AUTH] Login exitoso")
                return True
            except Exception:
                error_usuario = page.query_selector(
                    ".alert.alert-danger:has-text('No se ha encontrado un usuario registrado')"
                )
                error_pass = page.query_selector(
                    ".alert.alert-danger:has-text('La contraseña ingresada es incorrecta')"
                )
                if error_usuario or error_pass:
                    mensaje = "Usuario no encontrado" if error_usuario else "Contraseña incorrecta"
                    self.log(f"[AUTH] Error de autenticación: {mensaje}")
                else:
                    self.log("[AUTH] No se pudo confirmar el login.")
                return False

        except Exception as e:
            self.log(f"[AUTH] Error durante el proceso de login: {e}")
            return False

    def sincronizar_pedidos(self) -> Dict[str, Any]:
        """Sincroniza pedidos desde NODO usando un archivo temporal."""
        resultado = {"exito": False, "mensaje": "", "pedidos": None}
        temp_path = None

        with sync_playwright() as p:
            user_data_dir = os.path.join(tempfile.gettempdir(), "practical_pedidos_playwright_profile")
            browser = p.chromium.launch_persistent_context(
                user_data_dir=user_data_dir,
                headless=True,
                timeout=30000
            )
            page = browser.pages[0] if browser.pages else browser.new_page()

            try:
                target_url = f"{self.credenciales.servidor}/LibroVenta/Index"
                self.log(f"[PEDIDOS] Navegando a: {target_url}...")
                page.goto(target_url, timeout=45000, wait_until="domcontentloaded")

                is_login_page = "Login" in page.url or page.query_selector("#Correo") is not None
                if is_login_page:
                    self.log("[AUTH] Sesión no encontrada o expirada. Iniciando sesión...")
                    if not self._login_nodo(page):
                        resultado["mensaje"] = "Error de autenticación. Verifique usuario/contraseña."
                        return resultado
                    if "LibroVenta" not in page.url:
                        page.goto(target_url, timeout=45000, wait_until="domcontentloaded")
                else:
                    self.log("[AUTH] Sesión activa detectada. Omitiendo inicio de sesión.")

                page.wait_for_selector("#cboConcepto", timeout=20000)

                self.log("[PEDIDOS] Aplicando filtros...")
                page.select_option("#cboConcepto", "4")
                page.evaluate("$('#cboConcepto').trigger('change')")
                page.wait_for_timeout(1000)
                page.select_option("#cboConsolidado", "1")

                self.log("[PEDIDOS] Ejecutando consulta...")
                page.click("#btnBuscar")
                page.wait_for_load_state("networkidle")

                with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_file:
                    temp_path = tmp_file.name
                    self.log(f"[PEDIDOS] Archivo temporal: {temp_path}")

                self.log("[PEDIDOS] Descargando reporte...")
                with page.expect_download(timeout=60000) as dl:
                    page.click("#btnDownloadExcel")
                dl.value.save_as(temp_path)
                self.log("[PEDIDOS] Descarga completada")

                self.log("[PEDIDOS] Procesando datos...")
                wb = openpyxl.load_workbook(temp_path, read_only=True, data_only=True)
                sheet = wb.active

                headers = []
                for row in sheet.iter_rows(max_row=1, values_only=True):
                    headers = list(row)
                    break

                header_map = {name: idx for idx, name in enumerate(headers) if name is not None}
                tiene_fecha = 'Fecha' in header_map

                pedidos = []
                idx = 0
                for row_cells in sheet.iter_rows(min_row=2, values_only=True):
                    factura_idx = header_map.get('Factura')
                    if factura_idx is None or factura_idx >= len(row_cells) or row_cells[factura_idx] is None:
                        continue

                    factura = str(row_cells[factura_idx])

                    fecha_pedido = _FECHA_BASE + timedelta(seconds=idx)
                    if tiene_fecha:
                        fecha_idx = header_map['Fecha']
                        if fecha_idx < len(row_cells) and row_cells[fecha_idx] is not None:
                            fecha_val = row_cells[fecha_idx]
                            if isinstance(fecha_val, datetime):
                                fecha_pedido = fecha_val
                            else:
                                try:
                                    fecha_pedido = datetime.fromisoformat(str(fecha_val))
                                except Exception:
                                    try:
                                        fecha_pedido = datetime.strptime(str(fecha_val), '%Y-%m-%d %H:%M:%S')
                                    except Exception:
                                        pass

                    estado_idx = header_map.get('EstadoTransaccion')
                    estado = str(row_cells[estado_idx]) if (estado_idx is not None and estado_idx < len(row_cells) and row_cells[estado_idx] is not None) else ""

                    total_idx = header_map.get('Total')
                    total = 0.0
                    if total_idx is not None and total_idx < len(row_cells) and row_cells[total_idx] is not None:
                        try:
                            total = float(row_cells[total_idx])
                        except (ValueError, TypeError):
                            total = 0.0

                    cliente_idx = header_map.get('Cliente')
                    cliente = str(row_cells[cliente_idx]) if (cliente_idx is not None and cliente_idx < len(row_cells) and row_cells[cliente_idx] is not None) else ""

                    vendedor_idx = header_map.get('Vendedor')
                    vendedor = str(row_cells[vendedor_idx]) if (vendedor_idx is not None and vendedor_idx < len(row_cells) and row_cells[vendedor_idx] is not None) else ""

                    moneda_idx = header_map.get('Moneda')
                    moneda = str(row_cells[moneda_idx]) if (moneda_idx is not None and moneda_idx < len(row_cells) and row_cells[moneda_idx] is not None) else ""

                    pedido = Pedido(
                        numero_pedido=factura,
                        fecha=fecha_pedido,
                        estado=estado,
                        total=total,
                        cliente=cliente,
                        vendedor=vendedor,
                        moneda=moneda,
                    )
                    pedidos.append(pedido)
                    idx += 1

                wb.close()
                self.log(f"[PEDIDOS] Procesados {len(pedidos)} pedidos")

                resultado["exito"] = True
                resultado["mensaje"] = "Descarga y procesamiento exitoso"
                resultado["pedidos"] = pedidos

            except Exception as e:
                resultado["mensaje"] = f"Error: {str(e)}"
                self.log(f"[ERROR] {e}")
            finally:
                # RF-33: eliminar archivo temporal siempre
                if temp_path and os.path.exists(temp_path):
                    try:
                        os.unlink(temp_path)
                        self.log("[LIMPIAR] Archivo temporal eliminado.")
                    except Exception as e:
                        self.log(f"[LIMPIAR] Error al eliminar archivo temporal: {e}")

                browser.close()
                self.log("[PEDIDOS] Navegador cerrado")

        return resultado

    def probar_conexion(self) -> Dict[str, Any]:
        """Prueba de conexión sin descargar datos."""
        resultado = {"exito": False, "mensaje": ""}

        with sync_playwright() as p:
            user_data_dir = os.path.join(tempfile.gettempdir(), "practical_pedidos_playwright_profile")
            browser = p.chromium.launch_persistent_context(
                user_data_dir=user_data_dir,
                headless=True,
                timeout=30000
            )
            page = browser.pages[0] if browser.pages else browser.new_page()

            try:
                page.goto(self.credenciales.servidor, timeout=45000, wait_until="domcontentloaded")
                is_login_page = "Login" in page.url or page.query_selector("#Correo") is not None
                if is_login_page:
                    if self._login_nodo(page):
                        resultado["exito"] = True
                        resultado["mensaje"] = "Conexión exitosa"
                    else:
                        resultado["mensaje"] = "Error de autenticación. Verifique usuario/contraseña."
                else:
                    resultado["exito"] = True
                    resultado["mensaje"] = "Conexión exitosa (Sesión activa)"
            except Exception as e:
                resultado["mensaje"] = f"Error de conexión: {str(e)}"
            finally:
                browser.close()

        return resultado
